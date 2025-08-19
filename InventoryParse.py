from tqdm import tqdm
import openpyxl


def CreateSet(erp, wms):
    full_serial_set = set()
    for i in erp:
        full_serial_set.add(i.serial)   # extract serial only
    for j in wms:
        full_serial_set.add(j.serial)   # extract serial only
    return full_serial_set


class ExcelRow:
    def __init__(self, serial, inventory, device=""):
        self.serial = serial
        self.inventory = inventory
        self.device = device


def AnalyzeExcel(file_path, limit=None):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # skip header
        serial = row[0]
        inventory = row[1] if len(row) > 1 else ""
        device = row[2] if len(row) > 2 else ""

        if serial:  # skip blank serials
            rows.append(ExcelRow(serial, inventory, device))

    if limit:
        rows = rows[:limit]

    print(f"Processing {len(rows)} rows from {file_path}...")
    return rows   # <-- return objects, not strings!


class Serial:
    def __init__(self, serial):
        self.serial = serial
        self.WMS_location = ""
        self.ERP_location = ""
        self.device = ""
        self.status = ""

    def __str__(self):
        return f'Device: {self.device} | Serial: {self.serial} | ERP: {self.ERP_location} | WMS: {self.WMS_location} | Status: {self.status}'


def ConformInventoryNames(inventory_string):
    if not inventory_string:
        return ""
    inventory_string = inventory_string.upper()
    if "TRIAGE" in inventory_string:
        return "Triage"
    elif "RETAIL" in inventory_string:
        return "Retail"
    elif "SUB" in inventory_string:
        return "Sub-Wip"
    elif "QUAR" in inventory_string:
        return "Quar"
    elif "REPAIR" in inventory_string:
        return "Repair"
    else:
        return inventory_string


def Run():
    wms = AnalyzeExcel("./WMS.xlsx")
    erp = AnalyzeExcel("./ERP.xlsx")

    all_serials = list(CreateSet(erp, wms))
    serial_class_lst = [Serial(s) for s in all_serials]

    print(f"\nMatching {len(serial_class_lst)} serials...")
    for class_serial in tqdm(serial_class_lst, desc="Processing Serials"):
        for raw_serial_row in wms:
            if class_serial.serial == raw_serial_row.serial:
                class_serial.WMS_location = ConformInventoryNames(raw_serial_row.inventory)
                class_serial.device = raw_serial_row.device
                break

        for raw_serial_row in erp:
            if class_serial.serial == raw_serial_row.serial:
                class_serial.ERP_location = ConformInventoryNames(raw_serial_row.inventory)
                class_serial.device = raw_serial_row.device
                break

        class_serial.status = "Synced" if class_serial.ERP_location == class_serial.WMS_location else "Not Synced"

    # Print results to console (optional)
    print("\n=== Final Results ===")
    for final in serial_class_lst:
        print(final)

    # Write results to Excel
    WriteResultsToExcel(serial_class_lst, output_path="Serial_Report.xlsx")


def WriteResultsToExcel(serial_class_lst, output_path="Output.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Serial Report"

    # Add headers
    headers = ["Device", "Serial", "ERP Location", "WMS Location", "Status"]
    ws.append(headers)

    # Add data
    for s in serial_class_lst:
        ws.append([s.device, s.serial, s.ERP_location, s.WMS_location, s.status])

    wb.save(output_path)
    print(f"Results saved to {output_path}")


if __name__ == "__main__":
    Run()

